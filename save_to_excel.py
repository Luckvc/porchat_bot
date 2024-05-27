from openpyxl import load_workbook
from openpyxl.styles import Alignment


def append_to_spreadsheed(contacts):
    file_path = 'Leads New.xlsx'
    workbook = load_workbook(filename=file_path)
    worksheet = workbook.active

    first_empty_row = worksheet.max_row + 2
    center_alignment = Alignment(horizontal='center', vertical='center')

    for i in range(len(contacts[0])):
        name_cell = worksheet.cell(row=first_empty_row + i, column=1, value=contacts[0][i])
        phone_cell = worksheet.cell(row=first_empty_row + i, column=2, value=contacts[1][i])
        name_cell.alignment = center_alignment
        phone_cell.alignment = center_alignment

    workbook.save(filename=file_path)
